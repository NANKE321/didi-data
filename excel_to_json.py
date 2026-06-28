#!/usr/bin/env python3
"""
Excel → JSON 转换器（支持 .xls 和 .xlsx）
"""
import json, sys, os
from datetime import datetime, timedelta

def safe_num(val, default=0):
    if val is None or val == '-' or val == '':
        return default
    try:
        return float(val)
    except (ValueError, TypeError):
        return default

def safe_str(val, default='-'):
    if val is None or val == '':
        return default
    return str(val).strip()

def excel_date(val):
    """Excel 日期序列号 → YYYY-MM-DD"""
    if isinstance(val, float) and val > 40000:
        d = datetime(1899, 12, 30) + timedelta(days=int(val))
        return d.strftime('%Y-%m-%d')
    elif isinstance(val, datetime):
        return val.strftime('%Y-%m-%d')
    return str(val)[:10]

def read_excel(path):
    """读取 Excel，返回 (列名列表, 数据行列表)"""
    ext = os.path.splitext(path)[1].lower()

    if ext == '.xls':
        import xlrd
        wb = xlrd.open_workbook(path)
        # 找数据最多的 sheet
        ws = max(wb.sheets(), key=lambda s: s.nrows)
        print(f"  使用 sheet: {ws.name}")
        headers = [ws.cell_value(0, c) for c in range(ws.ncols)]
        rows = []
        for r in range(1, ws.nrows):
            row = [ws.cell_value(r, c) for c in range(ws.ncols)]
            rows.append(tuple(row))
        return headers, rows
    else:
        import openpyxl
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        if '今日数据' in wb.sheetnames:
            ws = wb['今日数据']
        else:
            ws = max(wb.worksheets, key=lambda s: s.max_row)
        print(f"  使用 sheet: {ws.title}")
        all_rows = list(ws.iter_rows(values_only=True))
        headers = list(all_rows[0])
        rows = all_rows[1:]
        wb.close()
        return headers, rows

def excel_to_json(excel_path, output_path):
    headers, data_rows = read_excel(excel_path)
    print(f"  找到 {len(data_rows)} 条记录")

    # 建立列名索引
    col = {}
    for i, h in enumerate(headers):
        if h:
            col[str(h).strip()] = i

    results = []
    for row in data_rows:
        def get(name, default=None):
            idx = col.get(name)
            if idx is not None and idx < len(row):
                return row[idx]
            return default

        name = safe_str(get('司机姓名', get('姓名', '')))
        if not name or name == '-':
            continue

        # 日期
        date_val = get('取数日期', get('滴滴数据取值日期'))
        date_str = excel_date(date_val) if date_val else datetime.now().strftime('%Y-%m-%d')

        entered = '1' if safe_num(get('是否入围', 0)) == 1 else '0'
        tier = safe_str(get('档位', '-'), '-')
        if tier in ('None', ''):
            tier = '-'

        compliance_val = safe_num(get('是否车证合规', 0))
        compliance = '双证合规' if compliance_val == 1 else '不合规'

        # 合规类型字段（如果有）
        compliance_type = safe_str(get('合规类型', ''))
        if compliance_type and compliance_type != '-':
            compliance = compliance_type

        billing_time = safe_num(get('计费时长（剔除培训）', get('你总计费时长', 0)))
        threshold = safe_num(get('入围门槛', get('当前入围门槛', 0)))
        diff = round(billing_time - threshold, 2)

        # 按整月剩余天数计算每天最低计费
        from calendar import monthrange
        try:
            dt = datetime.strptime(date_str, '%Y-%m-%d')
            days_in_month = monthrange(dt.year, dt.month)[1]
            remaining_days = days_in_month - dt.day
        except:
            remaining_days = 4  # 默认
        gap = threshold - billing_time
        daily_min_billing = round(gap / remaining_days, 2) if gap > 0 and remaining_days > 0 else 0.0

        peak_ratio = safe_num(get('工作日高峰计费占比', 0))
        fast_ratio = safe_num(get('快优订单占比', 0))
        if peak_ratio <= 1:
            peak_ratio = round(peak_ratio * 100, 1)
        if fast_ratio <= 1:
            fast_ratio = round(fast_ratio * 100, 1)

        record = {
            "date": date_str,
            "name": name,
            "plate": safe_str(get('车牌号', '-')),
            "status": safe_str(get('司机状态', '其他'), '其他'),
            "entered": entered,
            "tier": tier,
            "compliance": compliance,
            "days_worked": int(safe_num(get('在职天数', get('本月绑车天数', 0)))),
            "daily_service_score": round(safe_num(get('日均服务分', 0)), 2),
            "total_orders": int(safe_num(get('完单数', get('本月完单数', 0)))),
            "total_flow": round(safe_num(get('司机基础流水', get('你总基础流水', 0))), 2),
            "total_billing_time": round(billing_time, 2),
            "online_time": round(safe_num(get('在线时间', 0)), 2),
            "peak_online_time": round(safe_num(get('高峰在线时间', 0)), 2),
            "threshold": round(threshold, 2),
            "diff": diff,
            "daily_min_billing": daily_min_billing,
            "early_peak": round(safe_num(get('早高峰计费', 0)), 2),
            "noon_peak": round(safe_num(get('午高峰计费', 0)), 2),
            "late_peak": round(safe_num(get('晚高峰计费', 0)), 2),
            "night_peak": round(safe_num(get('夜高峰计费', 0)), 2),
            "workday_peak_billing": round(safe_num(get('工作日高峰计费', 0)), 2),
            "workday_billing": round(safe_num(get('工作日计费', 0)), 2),
            "peak_ratio": peak_ratio,
            "fast_ratio": fast_ratio,
            "fast_count": int(safe_num(get('快优订单数', get('快优单量', 0)))),
            "total_orders2": int(safe_num(get('订单数', get('本月完单数', 0)))),
            "vehicle_owner": safe_str(get('车辆所有人', '-')),
            "company": safe_str(get('公司名称', '上海超阳汽车租赁有限公司'))
        }
        results.append(record)

    # === 合并历史数据，保留最近 7 天 ===
    d7_path = os.path.join(os.path.dirname(output_path), 'data_7d.json')
    all_data = list(results)
    if os.path.exists(d7_path):
        try:
            with open(d7_path, 'r', encoding='utf-8') as f:
                old_data = json.load(f)
            new_dates = set(r['date'] for r in results)
            for r in old_data:
                if r['date'] not in new_dates:
                    all_data.append(r)
            all_dates = sorted(set(r['date'] for r in all_data), reverse=True)
            keep_dates = set(all_dates[:7])
            all_data = [r for r in all_data if r['date'] in keep_dates]
            print(f"  📅 合并后: {len(all_data)} 条, 日期: {sorted(keep_dates, reverse=True)}")
        except Exception as e:
            print(f"  ⚠️ 合并失败: {e}，使用新数据")

    all_data.sort(key=lambda r: (r['date'], -r['total_flow']))

    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(all_data, f, ensure_ascii=False, indent=None)
    with open(d7_path, 'w', encoding='utf-8') as f:
        json.dump(all_data, f, ensure_ascii=False, indent=None)

    print(f"  ✅ 转换完成: {len(all_data)} 条记录")

if __name__ == '__main__':
    if len(sys.argv) < 3:
        print("用法: python3 excel_to_json.py <input.xls/xlsx> <output.json>")
        sys.exit(1)
    excel_to_json(sys.argv[1], sys.argv[2])
